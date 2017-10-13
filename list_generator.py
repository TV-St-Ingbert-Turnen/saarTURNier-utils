from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet import Worksheet
from saar_teams import *
import datetime
import csv
import os


class ListGenerator(object):
    def __init__(self, path):
        self._path = path

    def _get_path(self, filename):
        return os.path.join(self._path, filename)

    def generate(self, team_list):
        raise NotImplementedError()

    def write(self):
        raise NotImplementedError()

    def close(self):
        raise NotImplementedError()


class ExcelGenerator(ListGenerator):
    def __init__(self, path, prefix):
        super().__init__(path)
        self._wb = None
        self._prefix = prefix

    def generate(self, team_list):
        raise NotImplementedError()

    def write(self):
        now = datetime.datetime.now()
        filename = "{}_{}.xlsx".format(self._prefix, now.strftime("%Y-%m-%d"))
        self._wb.save(super()._get_path(filename))

    def close(self):
        self._wb.close()


class RefereeCsvGenerator(ListGenerator):
    def generate(self, team_list):
        for tid, team in enumerate(team_list):
            assert isinstance(team, SaarTeam)
            for referee in team.referees:
                self._referees.append([referee, team.name, tid + 1])

    def __init__(self, path):
        super().__init__(path)
        self._referees = []

    def write(self):
        with open(super()._get_path('referees.csv'), 'w', newline='\n', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_ALL)
            writer.writerows(self._referees)

    def close(self):
        pass


class ScoreSystemCsvGenerator(ListGenerator):
    def __init__(self, path):
        super().__init__(path)
        self._teams = []
        self._participants = []
        # participants.csv: "Vorname Name"; "[w|m]"; "tid"
        # teams.csv: "tid";"Name"

    def generate(self, team_list):
        for tid, team in enumerate(team_list):
            assert isinstance(team, SaarTeam)
            self._teams.append([tid + 1, team.name])
            for gymnast in team.gymnasts:
                g_name = "{} {}".format(gymnast.name, gymnast.surname)
                g_gender = 'm' if gymnast.gender == SaarGymnast.MALE else 'w'
                self._participants.append([g_name, g_gender, tid + 1])

    def write(self):
        with open(super()._get_path('teams.csv'), 'w', newline='\n', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_ALL)
            writer.writerows(self._teams)
        with open(super()._get_path('participants.csv'), 'w', newline='\n', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_ALL)
            writer.writerows(self._participants)

    def close(self):
        pass


class RefereeFormsGenerator(ExcelGenerator):
    def __init__(self, path):
        super().__init__(path, "Wertungsbogen")

    def generate(self, team_list: SaarTeamList):

        self._wb = load_workbook(filename="./templates/Wertungsbogen_Master.xlsx")
        master_ws = self._wb.active
        version_string = master_ws["F2"].value
        check_version(version_string)

        i = 0
        for team in team_list:
            assert isinstance(team, SaarTeam)
            for apparatus_f, apparatus_m in [("Boden", "Boden"), ("Sprung", "Sprung"), ("Stufenbarren", "Reck"),
                                             ("Balken", "Barren")]:
                title = "{}_{}_{}".format(team.name[:10], apparatus_f[:2], apparatus_m[:2])
                print(title)

                ws = self._wb.create_sheet(title=title)
                ws_copy = WorksheetCopy(master_ws, ws)
                ws_copy.copy_worksheet()

                # set smaller page margins
                assert isinstance(ws, Worksheet)
                ws.page_margins = PageMargins(.2, .2, .75, .75, .314, .314)

                # set contents
                ws["F1"] = apparatus_f
                ws["F18"] = apparatus_m
                ws["A2"] = team.name
                ws["A19"] = team.name

                offset = 4
                for num, gymnast in enumerate(team.get_gymnasts(SaarGymnast.FEMALE)):
                    row = offset + num
                    ws["A{}".format(row)] = num + 1
                    ws["B{}".format(row)] = gymnast.name
                    ws["C{}".format(row)] = gymnast.surname

                offset = 21
                for num, gymnast in enumerate(team.get_gymnasts(SaarGymnast.MALE)):
                    row = offset + num
                    ws["A{}".format(row)] = num + 1
                    ws["B{}".format(row)] = gymnast.name
                    ws["C{}".format(row)] = gymnast.surname

                i += 1

        self._wb.remove(self._wb["Master"])
