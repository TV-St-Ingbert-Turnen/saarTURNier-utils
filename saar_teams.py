import os
from openpyxl import load_workbook
from openpyxl.worksheet import Worksheet
from version import check_version


class SaarTeamList:
    def __init__(self, teams_file: str):
        if not os.path.exists(teams_file):
            raise FileNotFoundError('{} was not found. Please provide the full path.'.format(teams_file))
        if not teams_file.lower().endswith('.xlsx'):
            raise IOError(
                '{} is not supported. Only Excel (*.xlsx) files are supported: use the correct template "Teilnehmerliste.xlsx".'.format(
                    teams_file))
        teams_wb = load_workbook(teams_file)

        version_string = teams_wb.active["N4"].value
        check_version(version_string)

        self._teams_wb = teams_wb
        self.teams = self._load()
        self._teams_wb.close()

    def __iter__(self):
        return iter(self.teams)

    def _extract_gymnast(self, cells, gender):
        name, surname, year = tuple([x.value for x in cells[0]])
        if name is None and surname is None:
            return None
        if name is None:
            name = ""
        if surname is None:
            surname = ""
        if year is None:
            year = 0
        return SaarGymnast(name, surname, year, gender)

    def _load(self):

        ws = self._teams_wb.active
        assert isinstance(ws, Worksheet)

        saar_teams = []

        for row in ws.iter_rows(max_col=1):
            cell = row[0]
            is_str = cell.data_type is 's'
            if is_str and 'Verein #' in cell.value:
                print(cell.value)
                team_name = ws['B{}'.format(cell.row)].value
                team = SaarTeam(team_name)

                for r in range(cell.row + 4, cell.row + 16):
                    print('gymnasts from {}:{}'.format(r, r))
                    # female gymnasts
                    cells = ws['A{}:C{}'.format(r, r)]
                    team.add_gymnast(self._extract_gymnast(cells, SaarGymnast.FEMALE))

                    # male gymnasts
                    cells = ws['E{}:G{}'.format(r, r)]
                    team.add_gymnast(self._extract_gymnast(cells, SaarGymnast.MALE))

                for r in range(cell.row + 18, cell.row + 21):
                    print('referees from {}:{}'.format(r, r))
                    # female gymnasts
                    team.add_referee(ws['A{}'.format(r)].value)

                contact_cells = ws['F{}:F{}'.format(cell.row + 18, cell.row + 21)]
                contact_data = [x[0].value for x in contact_cells]
                team.add_contact_data(contact_data)

                if not (team.name is None and len(team.gymnasts) == 0):
                    saar_teams.append(team)

        for t in saar_teams:
            assert isinstance(t, SaarTeam)
            if t.is_empty:
                saar_teams.remove(t)

        return saar_teams


class SaarTeam:
    def __init__(self, name):
        self._name = name
        self._gymnasts = []
        self._referees = []
        self._contact_data = None

    def __str__(self):
        return self.name

    def add_gymnast(self, gymnast):
        if gymnast is None:
            return 0
        self._gymnasts.append(gymnast)
        return 1

    def add_referee(self, referee):
        if referee is None:
            return 0
        self._referees.append(referee)

    def add_contact_data(self, data):
        self._contact_data = data

    def get_gymnasts(self, gender=None):
        self._gymnasts.sort(key=lambda x: x.name)
        if gender is None:
            return self.gymnasts
        elif gender is SaarGymnast.FEMALE:
            return [g for g in self.gymnasts if g.gender == SaarGymnast.FEMALE]
        else:
            return [g for g in self.gymnasts if g.gender == SaarGymnast.MALE]

    @property
    def name(self):
        return self._name

    @property
    def gymnasts(self):
        if len(self._gymnasts) == 0:
            return []
        return self._gymnasts

    @property
    def referees(self):
        return self._referees

    @property
    def contact(self):
        return self._contact_data

    @property
    def is_empty(self):
        return len(self._gymnasts) == 0 and self._name is None


class SaarGymnast:
    MALE, FEMALE = 'm', 'f'

    def __init__(self, name, surname, year, gender):
        self._name = name
        self._surname = surname
        self._year = year
        self._gender = gender

    def __str__(self):
        return "{} {}".format(self._name, self._surname)

    @property
    def name(self):
        return self._name

    @property
    def surname(self):
        return self._surname

    @property
    def year(self):
        return self._year

    @property
    def gender(self):
        return self._gender
