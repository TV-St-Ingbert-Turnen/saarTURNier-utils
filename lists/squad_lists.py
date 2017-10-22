from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import Worksheet
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import PatternFill, Font, NamedStyle, Alignment
from openpyxl.styles.builtins import accent_6, accent_5, accent_4, accent_2
import csv
from saar_teams import *
from constants import APPARATUS
from .list_generator import ListGenerator, ExcelGenerator
import os
from copy import copy

class CompetitionPlanGenerator(ExcelGenerator):
    def __init__(self, path, num_squads):
        super().__init__(path, "Wettkampfplan")
        self._num_squads = num_squads

        self._sq_style = ['Accent1', 'Accent2', 'Accent3', 'Accent6']

        self._ref_style = ['Normal', 'Normal', 'Normal', 'Normal']
        # ['20 % - Accent3', '20 % - Accent3', '20 % - Accent3', '20 % - Accent3']
        self._2ref = [
            [1, 0, 0, 1],
            [2, 2, 0, 0],
            [0, 1, 1, 0],
            [0, 0, 2, 2],
        ]
        self._4ref = [
            [1, 1, 1, 1],
            [2, 2, 2, 2],
            [3, 3, 3, 3],
            [4, 4, 4, 4],
        ]

    def _bold_font(self, cell):
        new_font = copy(cell.font)
        new_font.bold = True
        cell.font = new_font

    def _italic_font(self, cell):
        new_font = copy(cell.font)
        new_font.italic = True
        cell.font = new_font

    def _center(self, cell):
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _set_style(self, cell, style):
        border = copy(cell.border)
        cell.style = style
        cell.border = border


    def generate(self, team_list: SaarTeamList):

        self._wb = load_workbook(filename=os.path.join(os.getcwd(), "templates/Wettkampfplan.xlsx"))

        ws = self._wb.active

        squads = team_list.get_squads(num_squads=self._num_squads)

        # generate simple squad list
        row_offset = 31
        for sq_id, sq in enumerate(squads):
            for team in sq:
                ws["A{}".format(row_offset)].value = sq_id + 1
                ws["B{}".format(row_offset)].value = team.name
                for squad_cell in ws["A{}:D{}".format(row_offset, row_offset)][0]:
                    self._set_style(squad_cell, self._sq_style[sq_id])
                    self._bold_font(squad_cell)
                row_offset += 1

        # generate referee assignment table
        # one referee-set per squad (male/female); 4 are pre-defined
        if len(squads) == 2:
            # remove additional cols for referees
            for row in ws.iter_rows(min_row=19, max_row=25, min_col=4, max_col=5):
                for squad_cell in row:
                    squad_cell.value = None
                    squad_cell.border = None
                    squad_cell.style = 'Normal'
            # copy referee legend left
            for row in ws.iter_rows(min_row=20, max_row=25, min_col=6, max_col=6):
                for squad_cell in row:
                    n_cell = ws.cell(row=squad_cell.row, column=squad_cell.col_idx-2)
                    n_cell.value = squad_cell.value
                    n_cell.font = copy(squad_cell.font)
                    squad_cell.border = None
                    squad_cell.value = None

        # generate competition plan
        num_rotations = 4
        row_offset = 7
        col_offset = 2
        for rotation in range(4):
            for sq_id, sq in enumerate(squads):
                apparatus = divmod(rotation + sq_id, num_rotations)[1]
                col = col_offset + rotation
                row = row_offset + apparatus * 2
                squad_cell = ws.cell(column=col, row=row)
                squad_cell.value = "Riege {}".format(sq_id + 1)
                self._set_style(squad_cell, self._sq_style[sq_id])
                self._bold_font(squad_cell)
                self._center(squad_cell)

                referee_cell = ws.cell(column=col, row=row+1)
                ref_grid = self._2ref if self._num_squads == 2 else self._4ref
                ref = ref_grid[apparatus][rotation] - 1
                referee_cell.value = "KaRi {}".format(ref + 1)
                self._set_style(referee_cell, self._ref_style[ref])
                self._italic_font(referee_cell)
                self._center(referee_cell)
