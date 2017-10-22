from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import Worksheet
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.page import PageMargins
import csv
from saar_teams import *
from constants import APPARATUS
from .list_generator import ListGenerator, ExcelGenerator


class RefereeFormsGenerator(ExcelGenerator):
    def __init__(self, path):
        super().__init__(path, "Wertungsbogen")

    def generate(self, team_list: SaarTeamList):

        self._wb = load_workbook(filename="templates/Wertungsbogen_Master.xlsx")
        assert isinstance(self._wb, Workbook)
        master_ws = self._wb.active
        version_string = master_ws["F2"].value
        check_version(version_string)

        i = 0
        for team in team_list:
            assert isinstance(team, SaarTeam)
            for apparatus_f, apparatus_m in APPARATUS:
                title = "{}_{}_{}".format(team.name[:10], apparatus_f[:2], apparatus_m[:2])

                ws = self._wb.create_sheet(title=title)
                ws_copy = WorksheetCopy(master_ws, ws)
                ws_copy.copy_worksheet()

                # set smaller page margins
                assert isinstance(ws, Worksheet)
                ws.page_margins = PageMargins(.2, .2, .75, .75, .314, .314)

                # set contents
                ws["F1"] = apparatus_f
                ws["F18"] = apparatus_m
                ws["A2"] = team.name
                ws["A19"] = team.name

                offset = 4
                for num, gymnast in enumerate(team.get_gymnasts(SaarGymnast.FEMALE)):
                    row = offset + num
                    ws["A{}".format(row)] = num + 1
                    ws["B{}".format(row)] = gymnast.name
                    ws["C{}".format(row)] = gymnast.surname

                offset = 21
                for num, gymnast in enumerate(team.get_gymnasts(SaarGymnast.MALE)):
                    row = offset + num
                    ws["A{}".format(row)] = num + 1
                    ws["B{}".format(row)] = gymnast.name
                    ws["C{}".format(row)] = gymnast.surname

                i += 1

        self._wb.remove(self._wb["Master"])


class RefereeCsvGenerator(ListGenerator):
    def generate(self, team_list):
        for tid, team in enumerate(team_list):
            assert isinstance(team, SaarTeam)
            for referee in team.referees:
                self._referees.append([referee, team.name, tid + 1])

    def __init__(self, path):
        super().__init__(path)
        self._referees = []

    def write(self):
        with open(super()._get_path('referees.csv'), 'w', newline='\n', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_ALL)
            writer.writerows(self._referees)

    def close(self):
        pass